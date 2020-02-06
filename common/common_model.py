#!/usr/bin/env python3

# Filename: common_model.py

"""This app is for the Amadeus automation tasks."""
import os
import json

from PyQt5.QtWidgets import QDialog, QFileDialog
from PyQt5 import QtCore, QtGui

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

class Common_Model():

	def __init__(self):
		
		self.key_format = None
		self.base_format = None
		self.validity_format = None
		self.output_format = None
		self.reason_format = None
		self.compare_format = None
		self.bold_format = None
		self.border_format = None
		self.field_format = None

	#======= EXCEL =================================

	def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
	                       truncate_sheet=False, 
	                       **to_excel_kwargs):
	    """
	    Append a DataFrame [df] to existing Excel file [filename]
	    into [sheet_name] Sheet.
	    If [filename] doesn't exist, then this function will create it.
	    
	    Parameters:
	    filename : File path or existing ExcelWriter
	                (Example: '/path/to/file.xlsx')
	    df : dataframe to save to workbook
	    sheet_name : Name of sheet which will contain DataFrame.
	                (default: 'Sheet1')
	    startrow : upper left cell row to dump data frame.
	                Per default (startrow=None) calculate the last row
	                in the existing DF and write to the next row...
	    truncate_sheet : truncate (remove and recreate) [sheet_name]
	                    before writing DataFrame to Excel file
	    to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
	                        [can be dictionary]
	    
	    Returns: None
	    """
	    
	    # ignore [engine] parameter if it was passed
	    if 'engine' in to_excel_kwargs:
	        to_excel_kwargs.pop('engine')
	    
	    writer = pd.ExcelWriter(filename, engine='openpyxl')
	    
	    try:
	        # try to open an existing workbook
	        writer.book = load_workbook(filename)
	    
	        # get the last row in the existing Excel sheet
	        # if it was not specified explicitly
	        if startrow is None and sheet_name in writer.book.sheetnames:
	            startrow = writer.book[sheet_name].max_row
	    
	        # truncate sheet
	        if truncate_sheet and sheet_name in writer.book.sheetnames:
	            # index of [sheet_name] sheet
	            idx = writer.book.sheetnames.index(sheet_name)
	            # remove [sheet_name]
	            writer.book.remove(writer.book.worksheets[idx])
	            # create an empty sheet [sheet_name] using old index
	            writer.book.create_sheet(sheet_name, idx)
	    
	        # copy existing sheets
	        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
	    except FileNotFoundError:
	        # file does not exist yet, we will create it
	        pass
	    
	    if startrow is None:
	        startrow = 0
	    
	    # write out the new sheet
	    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
	    
	    # save the workbook
	    writer.save()

	def get_col_widths(self, dataframe):
		return [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

	def _fileDialog(self, directory='', forOpen=True, fmt='', isFolder=False):
		options = QFileDialog.Options()
		options |= QFileDialog.DontUseNativeDialog
		options |= QFileDialog.DontUseCustomDirectoryIcons
		dialog = QFileDialog()
		dialog.setOptions(options)

		dialog.setFilter(dialog.filter() | QtCore.QDir.Hidden)

		# ARE WE TALKING ABOUT FILES OR FOLDERS
		if isFolder:
			dialog.setFileMode(QFileDialog.DirectoryOnly)
		else:
			dialog.setFileMode(QFileDialog.AnyFile)

		# OPENING OR SAVING
		dialog.setAcceptMode(QFileDialog.AcceptOpen) if forOpen else dialog.setAcceptMode(QFileDialog.AcceptSave)

		# SET FORMAT, IF SPECIFIED
		if fmt != '' and isFolder is False:
			dialog.setDefaultSuffix(fmt)
			dialog.setNameFilters([f'{fmt} (*.{fmt})'])

		# SET THE STARTING DIRECTORY
		if directory != '':
			dialog.setDirectory(str(directory))
		else:
			main_dir = os.getcwd()
			dialog.setDirectory(str(main_dir))


		if dialog.exec_() == QDialog.Accepted:
			path = dialog.selectedFiles()[0]  # returns a list

			return path
		else:
			if directory !='':
				return directory
			else:
				return ''

		dialog.destroy()


	#======= PRICING VALIDATION========================

	def set_workbook_formats(self, workbook):

		# Add custom formats on the workbook.
		self.key_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'fg_color': '#f09886',
				'border': 1})

		self.base_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'fg_color': '#f3f7d0',
				'border': 1})

		self.validity_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'font_color': 'white',
				'fg_color': '#37b859',
				'border': 1})

		self.output_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'fg_color': '#5082d9',
				'border': 1})

		self.reason_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'font_color': 'white',
				'fg_color': '#145bd9',
				'border': 1})

		self.compare_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'fg_color': '#c2c1be',
				'border': 1})

		self.bold_format = workbook.add_format({'bold': 1}) 

		self.border_format = workbook.add_format({'bottom':1, 'top':1, 'left':1, 'right':1})

		self.field_format = workbook.add_format({
				'bold': True,
				'align': 'center',
				'valign': 'vcenter',
				#'text_wrap': True,
				'font_color': 'white',
				'fg_color': '#9bbb59',
				'border': 1})

	#======= SETTINGS =================================

	def _load_settings(self, report, settings_data, src_folder_ui, output_folder_ui):

		src_folder = ''
		output_folder = ''

		try:
			src_folder = settings_data[report]["SourceFolder"]
			output_folder = settings_data[report]["OutputFolder"]
		except KeyError:
			pass
			
		src_folder_ui.setText(src_folder)
		output_folder_ui.setText(output_folder)

	def _save_settings(self, report, src_folder, output_folder):

		path = os.getcwd() +"\\resources\\reference"
		json_file = path + "\\settings.json"

		self.updateJsonFile(report, src_folder, output_folder, json_file)


	def updateJsonFile(self, report, src_folder, tgt_folder, json_file):

		jsonFile = open(json_file, "r") # Open the JSON file for reading
		data = json.load(jsonFile) # Read the JSON into the buffer
		jsonFile.close() # Close the JSON file

		## Working with buffered content
		data[report]["SourceFolder"] = src_folder
		data[report]["OutputFolder"]  = tgt_folder

		## Save our changes to JSON file
		jsonFile = open(json_file, "w+")
		jsonFile.write(json.dumps(data))
		jsonFile.close()