import pandas as pd
import numpy as np
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

pd.options.mode.chained_assignment = None


def get_col_widths(dataframe):
    return [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

def merge_scale_price_df(df_scales, df_price):

	merged_df = pd.merge(df_scales,df_price,how='inner',left_on=['Short_text_key','Partner_Number','HybrisProduct_ID','BillingMaterial_ID_ToBe','SCALE_ID_MAP'], right_on = ['short_text_key','partner_number','hybrisproduct_id','billingmaterial_id_tobe','scale_id'])

	return merged_df

def merge_minf_price_df(df_minf, df_price):

	merged_df = pd.merge(df_minf,df_price,how='inner',left_on=['Short_text_key','Partner_Number','HybrisProduct_ID','BillingMaterial_ID_ToBe'], right_on = ['short_text_key','partner_number','hybrisproduct_id','billingmaterial_id_tobe'])

	return merged_df

keyword = 'MCK2'
usage_type = 'ARBK'

fn = r'C:\Users\rampil.saavedra\Documents\Amadeus DS App\resources\Pricing Validation\SF_PreVal_Pricing_Details_MCK2_orig.csv'

df_csv = pd.DataFrame()
for chunk in pd.read_csv(fn, chunksize=2000, dtype=str):
    df_csv = pd.concat([df_csv, chunk], ignore_index=True)

#Filter dataframe

preval_price_df = df_csv[df_csv['USAGE_TYPE_CD'] == usage_type]
#preval_price_df = df_csv

#Fill Nan as ''
preval_price_df.fillna(value='',inplace=True)

#Initialize dataframe

preval_price_df.drop(['Contract_Type','USAGE_TYPE_CD'], axis=1, inplace=True)

preval_price_df.rename(columns={
						'Bill_To': 'Bill-To',
						'Requestor': 'Requester', 
						'Price': 'Price_Amount', 
						'Pricing_Bundle_ID': 'Bundle_ID',
						'Pricing_Contract_Type': 'Contract_Type', 
						'Pricing_Contract_SubType': 'Contract_Sub_Type',
						'Maximum_Take_Off_Weight_Range': 'MTOW_Range',
						'Stock_Type': 'Stock_Ticket_Type',
						'Public_Mode': 'Publication_Mode',
						'Special_Price_Element': 'Special_pricing_element',
						'Hotel_Property_City': 'Hotel_Property_&_City',
						'Original_Booking_Channel': 'Bkg_Channel',
						'Priodicity': 'Periodicity',
						'Coupon_Status':'Coupons_Status',
						'Flight_Range_Code': 'Flight_Range'}, inplace=True)

# get max column
# Add columns to dataframe

col_to_add = [('Distribution_Channel','Dist_Channel'),('Office_Country','Off_Country'),('Office_Owner','Off_Owner'),('Office_Group','Off_Group'),
			('Reach_Indicator','Reach_Ind'),('Material_Group','Old_Material_Group')]

col_idx=[{'Curr_Name': c1, 'Col_Idx': preval_price_df.columns.get_loc(c1), 'New_Name' : c2} for c1,c2 in col_to_add if c1 in preval_price_df]

for idx in col_idx:

	max_col = preval_price_df.columns.to_list()[-1].split('_')[1]
	new_name = idx['New_Name']
	new_col_idx = int(max_col) + 1	
	col_loc = idx['Col_Idx']
	
	preval_price_df[new_name] = preval_price_df.iloc[:, col_loc]
	preval_price_df['Input_'+str(new_col_idx)] = preval_price_df.iloc[:, col_loc + 1]
	preval_price_df['Compare_'+str(new_col_idx)] = preval_price_df.iloc[:, col_loc + 2]

# preval_price_df['Bkg_Channel'] = preval_price_df['Original_Booking_Channel']
# preval_price_df['Accp_channel'] = preval_price_df['Acceptance_Channel']
# preval_price_df['Prod_Channel'] = preval_price_df['Product_Channel']
# preval_price_df['Group_Bkgs'] = preval_price_df['Group_Booking']
# preval_price_df['Mkt_Carrier'] = preval_price_df['Marketing_Carrier']
# preval_price_df['Flight_Range_Code'] = preval_price_df['Flight_Range']

#Read the template to get the column mapping

workbook = load_workbook(r'C:\Users\rampil.saavedra\Documents\Amadeus DS App\resources\Pricing Templates\Templates\OTC - 02 - ARBK - Air Booking - Usage type pricing template.xlsx', read_only = True, data_only=True)
price_ws = workbook['Customer prices']

rows = price_ws.iter_rows(min_row=2, max_row=2)
headers = next(rows)

template_columns = [(c.coordinate, c.value) for c in headers if 'last upd' not in c.value.lower() and 'created' not in c.value.lower()][4:]

workbook.close()

#Get the First 8 columns as Key Fields

price_df = preval_price_df.iloc[ : , : 9]

#Set dataframe column names to lower case

preval_price_df.columns = preval_price_df.columns.str.lower()
preval_price_df = preval_price_df.astype({"start_date": str, "end_date": str})

for col_idx, column in enumerate(template_columns):

	cell, col = column
	column_nm = '_'.join(col.split(' ')).lower()

	try:

		col_loc = preval_price_df.columns.get_loc(column_nm)

		# if column_nm in 'start_date':

		# 	price_df[col] = preval_price_df[['start_date']].apply(lambda x: '\'' + x['start_date'].split('.')[0] if x['start_date'] else x, axis=1)
		# 	price_df['Input_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 1].apply(lambda x: '\'' + x if x else x)
		# 	price_df['Compare_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 2].apply(lambda x: '\'' + x if x else x)

		# elif column_nm in 'end_date':

		# 	price_df[col] = preval_price_df[['end_date']].apply(lambda x: '\'' + x['end_date'].split('.')[0] if x['end_date'] else x, axis=1)
		# 	price_df['Input_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 1].apply(lambda x: '\'' + x if x else x)
		# 	price_df['Compare_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 2].apply(lambda x: '\'' + x if x else x)

		#else:

		price_df[col] = preval_price_df.iloc[:, col_loc]
		price_df['Input_'+str(col_idx+1)] = preval_price_df.iloc[:, col_loc + 1]
		price_df['Compare_'+str(col_idx+1)] = preval_price_df.iloc[:, col_loc + 2]

	except KeyError:

		print('{} is not existing on the dataframe'.format(column_nm))

	except Exception as e:

		print('Unknown error: ' + str(e))

#Write Header to Excel

wb_name = 'OTC_{}_Pricing_Pre_Validation_Report_{}.xlsx'.format(usage_type, keyword)

# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter(wb_name, engine='xlsxwriter', options={'strings_to_numbers': False})

# Get the xlsxwriter workbook and worksheet objects.
workbook  = writer.book

# Add custom formats on the workbook.
key_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'fg_color': '#f09886',
		'border': 1})

base_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'fg_color': '#f3f7d0',
		'border': 1})

validity_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'font_color': 'white',
		'fg_color': '#37b859',
		'border': 1})

output_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'fg_color': '#5082d9',
		'border': 1})

reason_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'font_color': 'white',
		'fg_color': '#145bd9',
		'border': 1})

compare_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'fg_color': '#c2c1be',
		'border': 1})

bold = workbook.add_format({'bold': 1}) 

border_fmt = workbook.add_format({'bottom':1, 'top':1, 'left':1, 'right':1})

# Summary Chart
 
ws_summary = workbook.add_worksheet('Pricing_Summary') 

match_cnt = len(price_df[price_df['Status']=='MATCHED'])
partial_match_cnt = len(price_df[price_df['Status']=='PARTIALLY MATCHED'])
not_found_cnt = len(price_df[price_df['Status']=='NOT FOUND'])

# Add the worksheet data that the charts will refer to. 
headings = ['Category', 'Values'] 
data = [ 
    ['Loaded Match', 'Loaded Partially Matched', 'Not Found'], 
    [match_cnt, partial_match_cnt, not_found_cnt], 
] 

# with bold format . 
ws_summary.write_row('B3', headings, bold) 
  
# Write a column of data starting from   
# 'A2', 'B2' respectively .  
ws_summary.write_column('B4', data[0]) 
ws_summary.write_column('C4', data[1]) 
 
# Create a chart object that can be added  
# to a worksheet using add_chart() method.  
  
# here we create a doughnut chart object .  
chart1 = workbook.add_chart({'type': 'doughnut'}) 
  
# Add a data series to a chart  
# using add_series method. 
  
# Configure the first series. 
# syntax to define ranges  
# [sheetname, first_row, first_col, last_row, last_col]. 

sheet_name = ws_summary.name
chart1.add_series({ 
    'name': 'Record Counts', 
 	'categories': [sheet_name, 3, 1, 5, 1], 
    'values':     [sheet_name, 3, 2, 5, 2], 
}) 

# Add a chart title  
chart1.set_title({'name': usage_type + ' - Record Counts'}) 
  
# Set an Excel chart style. Colors 
# with white outline and shadow. 
chart1.set_style(10) 
  
# add chart to the worksheet with an offset,  
# at the top-left corner of a chart   
# is anchored to cell C2 . 
ws_summary.insert_chart('E3', chart1, {'x_offset': 25, 'y_offset': 10}) 

# Field Comparison

# Add a header format.
field_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'font_color': 'white',
		'fg_color': '#9bbb59',
		'border': 1})

headings = ['Field Comparison', 'Matched', 'Not Matched', 'Total Count', 'Comments'] 
ws_summary.write_row('B20', headings, field_format) 

#Set header row heigth and width

ws_summary.set_column(1, 1, 22)
ws_summary.set_column(3, 5, 12)
ws_summary.set_row(19, 30)

l_price_df_cols = price_df.columns.to_list()
l_price_cols = l_price_df_cols[9::3]
l_price_fields = []

for col in l_price_cols:
	
	col_idx = l_price_df_cols.index(col)
	col_compare = l_price_df_cols.__getitem__(col_idx + 2)

	match_cnt = len(price_df[price_df[col_compare]=='Y'])
	no_match_cnt = len(price_df[price_df[col_compare]=='N'])
	total_cnt = match_cnt + no_match_cnt

	l_price_fields.append((col, match_cnt, no_match_cnt, total_cnt))

ws_summary.write_column('B21', [f for f,m,n,t in l_price_fields],border_fmt) 
ws_summary.write_column('C21', [m for f,m,n,t in l_price_fields],border_fmt) 
ws_summary.write_column('D21', [n for f,m,n,t in l_price_fields],border_fmt) 
ws_summary.write_column('E21', [t for f,m,n,t in l_price_fields],border_fmt)
ws_summary.write_column('F21', ['' for f,m,n,t in l_price_fields],border_fmt)

'''
Price Details
'''

sheet_name = 'Pricing_PreValidation_Details'
price_df.to_excel(writer, sheet_name=sheet_name, startrow=2, header=False, index=False)

worksheet = writer.sheets[sheet_name]

outputs_idx = [i for i,v in enumerate(price_df.columns.values) if v[:3] == 'End'][0]

# Write the column headers with the defined format.
for col_num, value in enumerate(price_df.columns.values):

	if col_num == 0:
		worksheet.write(1, col_num, value, reason_format)
	elif col_num == 1 or value[:5] == 'Start' or value[:3] == 'End':
		worksheet.write(1, col_num, value, validity_format)
	elif value[:5] == 'Input' or value[:7] == 'Compare':
		worksheet.write(1, col_num, value, compare_format)
	elif col_num > 1 and col_num < 9:
		worksheet.write(1, col_num, value, key_format)
	elif col_num > outputs_idx:
		worksheet.write(1, col_num, value, output_format)
	else:
		worksheet.write(1, col_num, value, base_format)

#Adjust the column width based on max values
for idx, width in enumerate(get_col_widths(price_df)):

	if idx in [0,2]:
		worksheet.set_column(idx, idx, 45)
	else:
		worksheet.set_column(idx, idx, width + 1)

#Adjust header height

worksheet.set_row(1, 30)

# Add borders

worksheet.conditional_format(xlsxwriter.utility.xl_range(1, 0, len(price_df) + 1, len(price_df.columns) - 1), {'type': 'no_errors', 'format': border_fmt})

# Add Header

worksheet.merge_range(0, 9, 0, outputs_idx - 4, 'INPUTS', base_format)
worksheet.merge_range(0, outputs_idx - 3, 0, outputs_idx + 2, 'VALIDITY', validity_format)
worksheet.merge_range(0, outputs_idx + 3, 0, len(price_df.columns) - 1, 'OUTPUTS', output_format)

# Add Filter

worksheet.autofilter(1,0,1,len(price_df.columns) - 1)

######################## SCALES ############################

fn = r'C:\Users\rampil.saavedra\Documents\Amadeus DS App\resources\Pricing Validation\SF_PreVal_Scales_Details_MCK2.csv'

df_csv = pd.DataFrame()
for chunk in pd.read_csv(fn, chunksize=2000, dtype=str):
    df_csv = pd.concat([df_csv, chunk], ignore_index=True)


preval_scales_df = df_csv[df_csv['USAGE_TYPE_CD'] == usage_type]

#Fill Nan as ''
preval_scales_df.fillna(value='',inplace=True)

#Initialize dataframe

preval_scales_df.drop(['USAGE_TYPE_CD'], axis=1, inplace=True)

#Read the template to get the column mapping

workbook = load_workbook(r'C:\Users\rampil.saavedra\Documents\Amadeus DS App\resources\Pricing Templates\Templates\OTC - 02 - ARBK - Air Booking - Usage type pricing template.xlsx', read_only = True, data_only=True)
price_ws = workbook['Customer scale price']

rows = price_ws.iter_rows(min_row=2, max_row=2)
headers = next(rows)

template_columns = [(c.coordinate, c.value) for c in headers if 'last upd' not in c.value.lower() and 'created' not in c.value.lower()][4:]

workbook.close()

merged_price_scale_df = merge_scale_price_df(preval_scales_df, preval_price_df)

#Get the First 8 columns as Key Fields

scales_df = merged_price_scale_df.iloc[ : , : 8]

merged_price_scale_df.columns = merged_price_scale_df.columns.str.lower()

for col_idx, column in enumerate(template_columns):

	cell, col = column
	column_nm = '_'.join(col.split(' ')).lower()

	try:

		col_loc = merged_price_scale_df.columns.get_loc(column_nm)

		scales_df[col] = merged_price_scale_df.iloc[:, col_loc]
		scales_df['Input_'+str(col_idx+1)] = merged_price_scale_df.iloc[:, col_loc + 1]
		scales_df['Compare_'+str(col_idx+1)] = merged_price_scale_df.iloc[:, col_loc + 2]

	except KeyError:

		print('{} is not existing on the Scales dataframe'.format(column_nm))

	except Exception as e:

		print('Unknown error: ' + str(e))

# Summary Chart

workbook  = writer.book
ws_summary = workbook.add_worksheet('Scales_Summary') 

match_cnt = len(scales_df[scales_df['Status']=='MATCHED'])
partial_match_cnt = len(scales_df[scales_df['Status']=='PARTIALLY MATCHED'])
not_found_cnt = len(scales_df[scales_df['Status']=='NOT FOUND'])

# Add the worksheet data that the charts will refer to. 
headings = ['Category', 'Values'] 
data = [ 
    ['Loaded Match', 'Loaded Partially Matched', 'Not Found'], 
    [match_cnt, partial_match_cnt, not_found_cnt], 
] 

# here we create bold format object .  
bold = workbook.add_format({'bold': 1}) 

# with bold format . 
ws_summary.write_row('B3', headings, bold) 
  
# Write a column of data starting from   
# 'A2', 'B2' respectively .  
ws_summary.write_column('B4', data[0]) 
ws_summary.write_column('C4', data[1]) 
 
# Create a chart object that can be added  
# to a worksheet using add_chart() method.  
  
# here we create a doughnut chart object .  
chart1 = workbook.add_chart({'type': 'doughnut'}) 

# Add a data series to a chart  
# using add_series method. 
  
# Configure the first series. 
# syntax to define ranges  
# [sheetname, first_row, first_col, last_row, last_col]. 

sheet_name = ws_summary.name
chart1.add_series({ 
    'name': 'Record Counts', 
 	'categories': [sheet_name, 3, 1, 5, 1], 
    'values':     [sheet_name, 3, 2, 5, 2], 
}) 

# Add a chart title  
chart1.set_title({'name': usage_type + ' - Scales Record Counts'}) 
  
# Set an Excel chart style. Colors 
# with white outline and shadow. 
chart1.set_style(10) 
  
# add chart to the worksheet with an offset,  
# at the top-left corner of a chart   
# is anchored to cell C2 . 
ws_summary.insert_chart('E3', chart1, {'x_offset': 25, 'y_offset': 10}) 

# Field Comparison

# Add a header format.
field_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'font_color': 'white',
		'fg_color': '#9bbb59',
		'border': 1})

headings = ['Field Comparison', 'Matched', 'Not Matched', 'Total Count', 'Comments'] 
ws_summary.write_row('B20', headings, field_format) 

#Set header row heigth and width

ws_summary.set_column(1, 1, 22)
ws_summary.set_column(3, 5, 12)
ws_summary.set_row(19, 30)

l_scales_df_cols = scales_df.columns.to_list()
l_scales_cols = l_scales_df_cols[8::3]
l_scales_fields = []

for col in l_scales_cols:
	
	col_idx = l_scales_df_cols.index(col)
	col_compare = l_scales_df_cols.__getitem__(col_idx + 2)

	match_cnt = len(scales_df[scales_df[col_compare]=='Y'])
	no_match_cnt = len(scales_df[scales_df[col_compare]=='N'])
	total_cnt = match_cnt + no_match_cnt

	l_scales_fields.append((col, match_cnt, no_match_cnt, total_cnt))

ws_summary.write_column('B21', [f for f,m,n,t in l_scales_fields],border_fmt) 
ws_summary.write_column('C21', [m for f,m,n,t in l_scales_fields],border_fmt) 
ws_summary.write_column('D21', [n for f,m,n,t in l_scales_fields],border_fmt) 
ws_summary.write_column('E21', [t for f,m,n,t in l_scales_fields],border_fmt)
ws_summary.write_column('F21', ['' for f,m,n,t in l_scales_fields],border_fmt)

'''
Scales Details
'''

sheet_name = 'Scales_PreValidation_Details'
scales_df.to_excel(writer, sheet_name=sheet_name, startrow=2, header=False, index=False)

worksheet = writer.sheets[sheet_name]

outputs_idx = [i for i,v in enumerate(scales_df.columns.values) if v[:3] == 'End'][0]

# Write the column headers with the defined format.
for col_num, value in enumerate(scales_df.columns.values):

	if col_num == 0:
		worksheet.write(1, col_num, value, reason_format)
	elif col_num == 1 or value[:5] == 'Start' or value[:3] == 'End':
		worksheet.write(1, col_num, value, validity_format)
	elif value[:5] == 'Input' or value[:7] == 'Compare':
		worksheet.write(1, col_num, value, compare_format)
	elif col_num > 1 and col_num < 8:
		worksheet.write(1, col_num, value, key_format)
	elif col_num > outputs_idx:
		worksheet.write(1, col_num, value, output_format)
	else:
		worksheet.write(1, col_num, value, base_format)

#Adjust the column width based on max values
for idx, width in enumerate(get_col_widths(scales_df)):

	if idx in [0,2]:
		worksheet.set_column(idx, idx, 45)
	else:
		worksheet.set_column(idx, idx, width + 1)

#Adjust header height

worksheet.set_row(1, 30)

# Add borders

worksheet.conditional_format(xlsxwriter.utility.xl_range(1, 0, len(scales_df) + 1, len(scales_df.columns) - 1), {'type': 'no_errors', 'format': border_fmt})

# Add Header

worksheet.merge_range(0, 8, 0, outputs_idx - 4, 'INPUTS', base_format)
worksheet.merge_range(0, outputs_idx - 3, 0, outputs_idx + 2, 'VALIDITY', validity_format)
worksheet.merge_range(0, outputs_idx + 3, 0, len(scales_df.columns) - 1, 'OUTPUTS', output_format)

# Add Filter

worksheet.autofilter(1,0,1,len(scales_df.columns) - 1)

######################## MINF ############################

fn = r'C:\Users\rampil.saavedra\Documents\Amadeus DS App\resources\Pricing Validation\SF_PreVal_MINF_Details_MCK2.csv'

df_csv = pd.DataFrame()
for chunk in pd.read_csv(fn, chunksize=2000, dtype=str):
    df_csv = pd.concat([df_csv, chunk], ignore_index=True)


#preval_minf_df = df_csv[df_csv['USAGE_TYPE_CD'] == usage_type]
preval_minf_df = df_csv

#Fill Nan as ''
preval_minf_df.fillna(value='',inplace=True)

#Initialize dataframe

preval_minf_df.drop(['USAGE_TYPE_CD'], axis=1, inplace=True)

#Read the template to get the column mapping

workbook = load_workbook(r'C:\Users\rampil.saavedra\Documents\Amadeus DS App\resources\Pricing Templates\Templates\OTC - 02 - ARBK - Air Booking - Usage type pricing template.xlsx', read_only = True, data_only=True)
price_ws = workbook['MINF Details MTable']

rows = price_ws.iter_rows(min_row=2, max_row=2)
headers = next(rows)

template_columns = [(c.coordinate, c.value) for c in headers if 'last upd' not in c.value.lower() and 'created' not in c.value.lower()][1:]

workbook.close()

#merged_price_minf_df = merge_minf_price_df(preval_minf_df, preval_price_df)

#Get the First 8 columns as Key Fields

minf_df = preval_minf_df.iloc[ : , : 6]

minf_df.columns = minf_df.columns.str.lower()

for col_idx, column in enumerate(template_columns):

	cell, col = column
	column_nm = '_'.join(col.split(' ')).lower()

	try:

		if column_nm in preval_minf_df.columns:
			
			col_loc = preval_minf_df.columns.get_loc(column_nm)
			
			minf_df[col] = preval_minf_df.iloc[:, col_loc]
			minf_df['Input_'+str(col_idx+1)] = preval_minf_df.iloc[:, col_loc + 1]
			minf_df['Compare_'+str(col_idx+1)] = preval_minf_df.iloc[:, col_loc + 2]

		else:

			col_loc = preval_price_df.columns.get_loc(column_nm)
		
			minf_df[col] = preval_price_df.iloc[:, col_loc]
			minf_df['Input_'+str(col_idx+1)] = preval_price_df.iloc[:, col_loc + 1]
			minf_df['Compare_'+str(col_idx+1)] = preval_price_df.iloc[:, col_loc + 2]

	except KeyError:

		print('{} is not existing on the MINF dataframe'.format(column_nm))

	except Exception as e:

		print('Unknown error: ' + str(e))


# Summary Chart

workbook  = writer.book
ws_summary = workbook.add_worksheet('MINF_Summary') 

match_cnt = len(minf_df[minf_df['status']=='MATCHED'])
partial_match_cnt = len(minf_df[minf_df['status']=='PARTIALLY MATCHED'])
not_found_cnt = len(minf_df[minf_df['status']=='NOT FOUND'])

# Add the worksheet data that the charts will refer to. 
headings = ['Category', 'Values'] 
data = [ 
    ['Loaded Match', 'Loaded Partially Matched', 'Not Found'], 
    [match_cnt, partial_match_cnt, not_found_cnt], 
]

# here we create bold format object .  
bold = workbook.add_format({'bold': 1}) 

# with bold format . 
ws_summary.write_row('B3', headings, bold) 
  
# Write a column of data starting from   
# 'A2', 'B2' respectively .  
ws_summary.write_column('B4', data[0]) 
ws_summary.write_column('C4', data[1]) 
 
# Create a chart object that can be added  
# to a worksheet using add_chart() method.  
  
# here we create a doughnut chart object .  
chart1 = workbook.add_chart({'type': 'doughnut'}) 

# Add a data series to a chart  
# using add_series method. 
  
# Configure the first series. 
# syntax to define ranges  
# [sheetname, first_row, first_col, last_row, last_col]. 

sheet_name = ws_summary.name
chart1.add_series({ 
    'name': 'Record Counts', 
 	'categories': [sheet_name, 3, 1, 5, 1], 
    'values':     [sheet_name, 3, 2, 5, 2], 
}) 

# Add a chart title  
chart1.set_title({'name': usage_type + ' - MINF Record Counts'}) 
  
# Set an Excel chart style. Colors 
# with white outline and shadow. 
chart1.set_style(10) 
  
# add chart to the worksheet with an offset,  
# at the top-left corner of a chart   
# is anchored to cell C2 . 
ws_summary.insert_chart('E3', chart1, {'x_offset': 25, 'y_offset': 10}) 

# Field Comparison

# Add a header format.
field_format = workbook.add_format({
		'bold': True,
		'align': 'center',
		'valign': 'vcenter',
		#'text_wrap': True,
		'font_color': 'white',
		'fg_color': '#9bbb59',
		'border': 1})

headings = ['Field Comparison', 'Matched', 'Not Matched', 'Total Count', 'Comments'] 
ws_summary.write_row('B20', headings, field_format) 

#Set header row heigth and width

ws_summary.set_column(1, 1, 22)
ws_summary.set_column(3, 5, 12)
ws_summary.set_row(19, 30)

l_minf_df_cols = minf_df.columns.to_list()
l_minf_cols = l_minf_df_cols[6::3]
l_minf_fields = []

for col in l_minf_cols:
	
	col_idx = l_minf_df_cols.index(col)
	col_compare = l_minf_df_cols.__getitem__(col_idx + 2)

	match_cnt = len(minf_df[minf_df[col_compare]=='Y'])
	no_match_cnt = len(minf_df[minf_df[col_compare]=='N'])
	total_cnt = match_cnt + no_match_cnt

	l_minf_fields.append((col, match_cnt, no_match_cnt, total_cnt))

ws_summary.write_column('B21', [f for f,m,n,t in l_minf_fields],border_fmt) 
ws_summary.write_column('C21', [m for f,m,n,t in l_minf_fields],border_fmt) 
ws_summary.write_column('D21', [n for f,m,n,t in l_minf_fields],border_fmt) 
ws_summary.write_column('E21', [t for f,m,n,t in l_minf_fields],border_fmt)
ws_summary.write_column('F21', ['' for f,m,n,t in l_minf_fields],border_fmt)

'''
MINF Details
'''

sheet_name = 'MINF_PreValidation_Details'
minf_df.to_excel(writer, sheet_name=sheet_name, startrow=2, header=False, index=False)

worksheet = writer.sheets[sheet_name]

print(minf_df.columns.values)

#outputs_idx = [i for i,v in enumerate(minf_df.columns.values) if v[:3] == 'End'][0]
outputs_idx = 1

# Write the column headers with the defined format.
for col_num, value in enumerate(minf_df.columns.values):

	if col_num == 0:
		worksheet.write(1, col_num, value, reason_format)
	elif col_num == 1 or value[:5] == 'Start' or value[:3] == 'End':
		worksheet.write(1, col_num, value, validity_format)
	elif value[:5] == 'Input' or value[:7] == 'Compare':
		worksheet.write(1, col_num, value, compare_format)
	elif col_num > 1 and col_num < 6:
		worksheet.write(1, col_num, value, key_format)
	elif col_num > outputs_idx:
		worksheet.write(1, col_num, value, output_format)
	else:
		worksheet.write(1, col_num, value, base_format)

#Adjust the column width based on max values
for idx, width in enumerate(get_col_widths(minf_df)):

	if idx in [0,2]:
		worksheet.set_column(idx, idx, 45)
	else:
		worksheet.set_column(idx, idx, width + 1)

#Adjust header height

worksheet.set_row(1, 30)

# Add borders

worksheet.conditional_format(xlsxwriter.utility.xl_range(1, 0, len(minf_df) + 1, len(minf_df.columns) - 1), {'type': 'no_errors', 'format': border_fmt})

# Add Header

#worksheet.merge_range(0, 6, 0, outputs_idx - 4, 'INPUTS', base_format)
#worksheet.merge_range(0, outputs_idx - 3, 0, outputs_idx + 2, 'VALIDITY', validity_format)
#worksheet.merge_range(0, outputs_idx + 3, 0, len(minf_df.columns) - 1, 'OUTPUTS', output_format)

# Add Filter

worksheet.autofilter(1,0,1,len(minf_df.columns) - 1)

# Close the Pandas Excel writer and output the Excel file.
writer.save()