#!/usr/bin/env python

import os
import pandas as pd
import re

from PyQt5.QtWidgets import QDialog, QMessageBox

import numpy as np
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from model.ui_model import main_app_Model

class Pricing_Validation_Model():

	def __init__(self, common_model, pt_model):

		self._model_ui = main_app_Model()
		self._common_mod = common_model
		self._pt_model = pt_model

		self._config = self._model_ui._conf_data['PrincingValConfig']

	#################### MAIN PROGRAM ##########################

	def read_source_to_DF(self, keyword, src_folder, pricing_tab):

		pd.options.mode.chained_assignment = None

		config = self._model_ui._conf_data[pricing_tab+'Config']

		src_filename = src_folder + "\\" +  config[pricing_tab+'ValFName'] + keyword + '.csv'
		
		df_csv = pd.DataFrame()

		for chunk in pd.read_csv(src_filename, chunksize=2000, dtype=str, encoding='utf-8', na_filter = False):
		    df_csv = pd.concat([df_csv, chunk], ignore_index=True)

		return df_csv

	def initialize_preval_df(self, preval_df, tab):

		#Fill Nan as ''
		preval_df.fillna(value='',inplace=True)

		#Initialize dataframe

		if tab == 'Price':

			preval_df.drop(['Contract_Type','USAGE_TYPE_CD'], axis=1, inplace=True)

			preval_df.rename(columns={
							'Bill_To': 'Bill-To',
							'Requestor': 'Requester', 
							'Price': 'Price_Amount', 
							'Pricing_Bundle_ID': 'Bundle_ID',
							'Pricing_Contract_Type': 'Contract_Type', 
							'Pricing_Contract_SubType': 'Contract_Sub_Type',
							'Maximum_Take_Off_Weight_Range': 'MTOW_Range',
							'Stock_Type': 'Stock_Ticket_Type',
							'Public_Mode': 'Publication_Mode',
							'Special_Price_Element': 'Special_pricing_element',
							'Hotel_Property_City': 'Hotel_Property_&_City',
							'Original_Booking_Channel': 'Bkg_Channel',
							'Priodicity': 'Periodicity',
							'Coupon_Status':'Coupons_Status',
							'Flight_Range_Code': 'Flight_Range'}, inplace=True)

			# get max column
			# Add columns to dataframe

			col_to_add = [('Distribution_Channel','Dist_Channel'),('Office_Country','Off_Country'),('Office_Owner','Off_Owner'),('Office_Group','Off_Group'),
						('Reach_Indicator','Reach_Ind'),('Material_Group','Old_Material_Group')]

			col_idx=[{'Curr_Name': c1, 'Col_Idx': preval_df.columns.get_loc(c1), 'New_Name' : c2} for c1,c2 in col_to_add if c1 in preval_df]

			for idx in col_idx:

				max_col = preval_df.columns.to_list()[-1].split('_')[1]
				new_name = idx['New_Name']
				new_col_idx = int(max_col) + 1	
				col_loc = idx['Col_Idx']
				
				preval_df[new_name] = preval_df.iloc[:, col_loc]
				preval_df['Input_'+str(new_col_idx)] = preval_df.iloc[:, col_loc + 1]
				preval_df['Compare_'+str(new_col_idx)] = preval_df.iloc[:, col_loc + 2]

			#Set dataframe column names to lower case

			preval_df = preval_df.astype({"Start_Date": str, "End_Date": str})

		else:

			preval_df.drop(['USAGE_TYPE_CD'], axis=1, inplace=True)


		return preval_df

	def get_template_mappings(self, usage_type, sheet_name, idx):

		#Read the template to get the column mapping

		config = self._model_ui._conf_data['PricingTempConfig']
		usageflow_conf = self._model_ui._conf_data['usageFlows']

		folder_path = os.getcwd() + "\\resources\\" + config['TemplateFolder']

		report_name = [usage_flow['ReportName'] for usage_flow in usageflow_conf if usage_flow['UsageType'] == usage_type][0]

		filename = folder_path + "\\" + report_name
		workbook = load_workbook(filename, read_only = True, data_only=True)
		ws = workbook[sheet_name]

		rows = ws.iter_rows(min_row=2, max_row=2)
		headers = next(rows)

		template_columns = [(c.coordinate, c.value) for c in headers if 'last upd' not in c.value.lower() and 'created' not in c.value.lower()][idx:]

		workbook.close()

		return template_columns

	def generate_template_df(self, usage_type, preval_df, sheet_nm, template_columns, idx, tab, price_df=None):

		l_nomapping = []

		#Get the Key Fields

		template_df = preval_df.iloc[ : , : idx]

		for col_idx, column in enumerate(template_columns):

			cell, col = column
			column_nm = '_'.join(col.split(' '))

			try:

				if tab == 'Price':

					col_loc = preval_df.columns.get_loc(column_nm)

					# if column_nm in 'start_date':

					# 	price_df[col] = preval_price_df[['start_date']].apply(lambda x: '\'' + x['start_date'].split('.')[0] if x['start_date'] else x, axis=1)
					# 	price_df['Input_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 1].apply(lambda x: '\'' + x if x else x)
					# 	price_df['Compare_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 2].apply(lambda x: '\'' + x if x else x)

					# elif column_nm in 'end_date':

					# 	price_df[col] = preval_price_df[['end_date']].apply(lambda x: '\'' + x['end_date'].split('.')[0] if x['end_date'] else x, axis=1)
					# 	price_df['Input_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 1].apply(lambda x: '\'' + x if x else x)
					# 	price_df['Compare_'+str(col_idx)] = preval_price_df.iloc[:, col_loc + 2].apply(lambda x: '\'' + x if x else x)

					#else:

					template_df[col] = preval_df.iloc[:, col_loc]
					template_df['Input_'+str(col_idx+1)] = preval_df.iloc[:, col_loc + 1]
					template_df['Compare_'+str(col_idx+1)] = preval_df.iloc[:, col_loc + 2]

				else:

					if column_nm in preval_df.columns:
						
						col_loc = preval_df.columns.get_loc(column_nm)

						template_df[col] = preval_df.iloc[:, col_loc]
						template_df['Input_'+str(col_idx+1)] = preval_df.iloc[:, col_loc + 1]
						template_df['Compare_'+str(col_idx+1)] = preval_df.iloc[:, col_loc + 2]

					else:

						col_loc = price_df.columns.get_loc(column_nm)
					
						template_df[col] = price_df.iloc[:, col_loc]
						template_df['Input_'+str(col_idx+1)] = price_df.iloc[:, col_loc + 1]
						template_df['Compare_'+str(col_idx+1)] = price_df.iloc[:, col_loc + 2]

			except KeyError:

				print('{} is not existing on the {} dataframe'.format(column_nm, tab))

				l_nomapping.append({'Sheet': sheet_nm, 'Template_Column': column_nm, 'Template_Cell' : 'NA'})

			except Exception as e:

				print('Unknown error: ' + str(e))


			#Append List no mapping to templates 

			self._pt_model.append_template_logs(usage_type, 2, tab, l_nomapping)

		return template_df


	def populate_summary_chart(self, usage_type, workbook, preval_df, sheet_nm, idx):
 
		ws_summary = workbook.add_worksheet(sheet_nm) 

		match_cnt = len(preval_df[preval_df['Status']=='MATCHED'])
		partial_match_cnt = len(preval_df[preval_df['Status']=='PARTIALLY MATCHED'])
		not_found_cnt = len(preval_df[preval_df['Status']=='NOT FOUND'])

		# Add the worksheet data that the charts will refer to. 
		headings = ['Category', 'Values'] 
		data = [ 
		    ['Loaded Match', 'Loaded Partially Matched', 'Not Found'], 
		    [match_cnt, partial_match_cnt, not_found_cnt], 
		] 

		# with bold format . 
		ws_summary.write_row('B3', headings, self._common_mod.bold_format) 
		  
		# Write the data to the columns
		ws_summary.write_column('B4', data[0]) 
		ws_summary.write_column('C4', data[1]) 
		 
		# Create a chart object that can be added  
		# to a worksheet using add_chart() method.  
		  
		# here we create a doughnut chart object .  
		chart1 = workbook.add_chart({'type': 'doughnut'}) 
		  
		# Add a data series to a chart  
		# using add_series method. 
		  
		# Configure the first series. 
		# syntax to define ranges  
		# [sheetname, first_row, first_col, last_row, last_col]. 

		sheet_name = ws_summary.name
		chart1.add_series({ 
		    'name': 'Record Counts', 
		 	'categories': [sheet_name, 3, 1, 5, 1], 
		    'values':     [sheet_name, 3, 2, 5, 2], 
		}) 

		# Add a chart title  
		chart1.set_title({'name': usage_type + ' - Record Counts'}) 
		  
		# Set an Excel chart style. Colors 
		# with white outline and shadow. 
		chart1.set_style(10) 
		  
		# add chart to the worksheet with an offset,  
		# at the top-left corner of a chart   
		# is anchored to cell C2 . 
		ws_summary.insert_chart('E3', chart1, {'x_offset': 25, 'y_offset': 10}) 

		#Set header row heigth and width

		ws_summary.set_column(1, 1, 22)
		ws_summary.set_column(3, 5, 12)
		ws_summary.set_row(19, 30)

		# Field Comparison

		if match_cnt != 0 or partial_match_cnt != 0 or not_found_cnt != 0:

			headings = ['Field Comparison', 'Matched', 'Not Matched', 'Total Count', 'Comments'] 
			ws_summary.write_row('B20', headings, self._common_mod.field_format) 

			l_preval_df_cols = preval_df.columns.to_list()
			l_preval_cols = l_preval_df_cols[idx::3]
			l_preval_fields = []

			for col in l_preval_cols:
				
				col_idx = l_preval_df_cols.index(col)
				col_compare = l_preval_df_cols.__getitem__(col_idx + 2)

				match_cnt = len(preval_df[preval_df[col_compare]=='Y'])
				no_match_cnt = len(preval_df[preval_df[col_compare]=='N'])
				total_cnt = match_cnt + no_match_cnt

				l_preval_fields.append((col, match_cnt, no_match_cnt, total_cnt))

			ws_summary.write_column('B21', [f for f,m,n,t in l_preval_fields],self._common_mod.border_format) 
			ws_summary.write_column('C21', [m for f,m,n,t in l_preval_fields],self._common_mod.border_format) 
			ws_summary.write_column('D21', [n for f,m,n,t in l_preval_fields],self._common_mod.border_format) 
			ws_summary.write_column('E21', [t for f,m,n,t in l_preval_fields],self._common_mod.border_format)
			ws_summary.write_column('F21', ['' for f,m,n,t in l_preval_fields],self._common_mod.border_format)

		
	def populate_details_worksheet(self, writer, preval_df, sheet_nm, merge_idx):

		preval_df.to_excel(writer, sheet_name=sheet_nm, startrow=2, header=False, index=False)

		worksheet = writer.sheets[sheet_nm]

		if sheet_nm != 'MINF_PreValidation_Details':
			outputs_idx = [i for i,v in enumerate(preval_df.columns.values) if v[:3] == 'End'][0]
		else:
			outputs_idx = 1

		# Write the column headers with the defined format.
		for col_num, value in enumerate(preval_df.columns.values):

			if col_num == 0:
				worksheet.write(1, col_num, value, self._common_mod.reason_format)
			elif col_num == 1 or value[:5] == 'Start' or value[:3] == 'End':
				worksheet.write(1, col_num, value, self._common_mod.validity_format)
			elif value[:5] == 'Input' or value[:7] == 'Compare':
				worksheet.write(1, col_num, value, self._common_mod.compare_format)
			elif col_num > 1 and col_num < merge_idx:
				worksheet.write(1, col_num, value, self._common_mod.key_format)
			elif col_num > outputs_idx:
				worksheet.write(1, col_num, value, self._common_mod.output_format)
			else:
				worksheet.write(1, col_num, value, self._common_mod.base_format)

		#Adjust the column width based on max values
		for idx, width in enumerate(self._common_mod.get_col_widths(preval_df)):

			if idx in [0,2]:
				worksheet.set_column(idx, idx, 45)
			else:
				worksheet.set_column(idx, idx, width + 1)

		#Adjust header height

		worksheet.set_row(1, 30)

		# Add borders

		worksheet.conditional_format(xlsxwriter.utility.xl_range(1, 0, len(preval_df) + 1, len(preval_df.columns) - 1), {'type': 'no_errors', 'format': self._common_mod.border_format})

		# Add Header

		if sheet_nm != 'MINF_PreValidation_Details':
			worksheet.merge_range(0, merge_idx, 0, outputs_idx - 4, 'INPUTS', self._common_mod.base_format)
			worksheet.merge_range(0, outputs_idx - 3, 0, outputs_idx + 2, 'VALIDITY', self._common_mod.validity_format)
			worksheet.merge_range(0, outputs_idx + 3, 0, len(preval_df.columns) - 1, 'OUTPUTS', self._common_mod.output_format)

		# Add Filter

		worksheet.autofilter(1,0,1,len(preval_df.columns) - 1)

	def merge_scale_price_df(self, df_scales, df_price):

		merged_df = pd.merge(df_scales,df_price,how='left',left_on=['Short_text_key','Partner_Number','HybrisProduct_ID','BillingMaterial_ID_ToBe','SCALE_ID_MAP'], right_on = ['Short_Text_Key','Partner_Number','HybrisProduct_ID','BillingMaterial_ID_ToBe','Scale_ID'])

		merged_df.rename(columns={'Status_x': 'Status','Reason_x':'Reason'}, inplace=True)

		return merged_df

	def merge_minf_price_df(self, df_minf, df_price):

		merged_df = pd.merge(df_minf,df_price,how='left',left_on=['Short_text_key','Partner_Number','HybrisProduct_ID','BillingMaterial_ID_ToBe','MF_Counter_ID'], right_on = ['Short_Text_Key','Partner_Number','HybrisProduct_ID','BillingMaterial_ID_ToBe','MF_Counter'])

		merged_df.rename(columns={'Status_x': 'Status','Reason_x':'Reason'}, inplace=True)
		
		return merged_df
